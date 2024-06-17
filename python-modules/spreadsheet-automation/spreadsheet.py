import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_inv_value_per_supplier = {}
products_inv_under_10 = {}


for product_row in range(2, product_list.max_row + 1):
    supplier_name =  product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    
    # Analysis-1: Number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # Analysis-2: Total value of inventory per supplier
    if supplier_name in total_inv_value_per_supplier:
        current_total_value = total_inv_value_per_supplier.get(supplier_name)
        total_inv_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_inv_value_per_supplier[supplier_name] = inventory * price

    # Analysis-3: Find the products wtih inventory numbers less than 10
    if inventory < 10:
        products_inv_under_10[int(product_num)] = int(inventory)

    # add value to spreadsheet for the total inventory price
    inventory_price.value = inventory * price


inventory_file.save("inventory_with_total_values.xlsx")
